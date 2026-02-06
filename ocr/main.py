from fastapi import FastAPI, UploadFile, File, HTTPException
from paddleocr import PaddleOCR
from pdf2image import convert_from_path
import docx
import openpyxl
import shutil
import os
import uvicorn
import logging

# Configure logging to suppress verbose Paddle output
logging.getLogger("ppocr").setLevel(logging.ERROR)

app = FastAPI()

# Initialize PaddleOCR
# 'use_angle_cls=True' enables the orientation classifier.
# Note: You might see a warning to use 'use_textline_orientation', 
# but use_angle_cls is still supported for initialization in v3.
ocr_engine = PaddleOCR(use_angle_cls=True, lang='en')

def extract_from_image(image_path):
    # FIX: Removed cls=True (deprecated/removed in newer PaddleOCR versions)
    # The engine uses the 'use_angle_cls' setting from initialization.
    result = ocr_engine.ocr(image_path)
    
    text = ""
    if result and result[0]:
        # Filter out None values just in case
        text = "\n".join([line[1][0] for line in result[0] if line and line[1]])
    return text

@app.get("/")
async def root():
    return {"status": "OCR Service is running"}

@app.post("/ocr")
async def perform_ocr(file: UploadFile = File(...)):
    filename = file.filename.lower()
    temp_file = f"temp_{file.filename}"
    
    # Save uploaded file
    with open(temp_file, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    full_text = ""

    try:
        # 1. Handle PDF
        if filename.endswith(".pdf"):
            # Convert PDF to images (200 DPI is usually enough and faster than 300)
            images = convert_from_path(temp_file, dpi=200)
            
            for i, image in enumerate(images):
                page_path = f"{temp_file}_page_{i}.jpg"
                image.save(page_path, "JPEG")
                
                page_text = extract_from_image(page_path)
                full_text += f"\n--- Page {i+1} ---\n{page_text}"
                
                if os.path.exists(page_path):
                    os.remove(page_path)

        # 2. Handle Word (.docx)
        elif filename.endswith(".docx"):
            doc = docx.Document(temp_file)
            full_text = "\n".join([para.text for para in doc.paragraphs])

        # 3. Handle Excel (.xlsx)
        elif filename.endswith(".xlsx"):
            wb = openpyxl.load_workbook(temp_file)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                full_text += f"\n--- Sheet: {sheet} ---\n"
                for row in ws.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) for cell in row if cell])
                    full_text += row_text + "\n"
                    
        # 4. Handle Images
        else:
            full_text = extract_from_image(temp_file)
        
        return {"text": full_text.strip()}
        
    except Exception as e:
        print(f"Error processing file: {e}")
        # Return a 500 error so the Laravel job knows it failed
        raise HTTPException(status_code=500, detail=str(e))
        
    finally:
        if os.path.exists(temp_file):
            os.remove(temp_file)

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)